import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_basis_sheet(filename="basis_recovery_sheet.xlsx"):
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Setup
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F2937")
    section_font = Font(name=font_family, size=12, bold=True, color="1F2937")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="374151")
    bold_data_font = Font(name=font_family, size=10, bold=True, color="111827")
    italic_help_font = Font(name=font_family, size=9, italic=True, color="6B7280")
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
    stripe_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Slate 50
    alert_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid") # Indigo 100
    
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    double_border_side = Side(border_style="double", color="94A3B8")
    
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(top=thin_border_side, bottom=double_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ----------------------------------------------------
    # Sheet 1: Instructions & Triage Checklist
    # ----------------------------------------------------
    ws_inst = wb.active
    ws_inst.title = "Triage & Instructions"
    ws_inst.views.sheetView[0].showGridLines = True
    
    ws_inst["A1"] = "Retirement Basis Recovery Triage Worksheet"
    ws_inst["A1"].font = title_font
    ws_inst.row_dimensions[1].height = 30
    
    ws_inst["A3"] = "Instructions:"
    ws_inst["A3"].font = section_font
    
    instructions = [
        "1. Use this workbook to manually reconstruct and track your retirement tax basis for state income tax purposes (NJ, PA, MA).",
        "2. Locate your historical contribution and distribution records using the Document Triage Checklist below.",
        "3. Select the appropriate tab for your state of residency (NJ/MA for pro-rata, PA for cost recovery/FIFO).",
        "4. Fill in the 'Year', 'Annual Contributions', 'Total Distributions', and 'Dec 31 Account Value' columns.",
        "5. The Excel formulas will automatically compute the 'Excludable' (tax-free) and 'Taxable' portions, and roll forward your 'Ending Basis' to the next year.",
        "6. Retain this spreadsheet and your supporting documents for your permanent tax files. Do not submit this to the state unless audited."
    ]
    
    for idx, inst in enumerate(instructions):
        row = 4 + idx
        ws_inst.cell(row=row, column=1, value=inst).font = data_font
        ws_inst.row_dimensions[row].height = 20
        
    ws_inst.cell(row=11, column=1, value="Document Triage Checklist:").font = section_font
    
    # Document Checklist Table Headers
    checklist_headers = ["Doc Type", "Key Target Data Field", "Where to Find It", "Triage Purpose"]
    for col_idx, h in enumerate(checklist_headers, start=1):
        cell = ws_inst.cell(row=13, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left
        cell.border = data_border
    ws_inst.row_dimensions[13].height = 25
    
    checklist_data = [
        ["Form W-2", "Box 16 (State Wages) vs Box 1 (Fed Wages)", "Compare State vs Federal wages on annual W-2 forms", "NJ/PA tax 403b/457 contributions. Box 16 - Box 1 = basis added."],
        ["Form 5498", "Box 1 (Traditional IRA contributions)", "Issued annually by the account custodian (Fidelity, Vanguard, etc.)", "Establishes total IRA contributions made for the year."],
        ["Form 1099-R", "Box 1 (Gross Distribution) & Box 7 (Code)", "Issued by custodian for any withdrawal during the year", "Determines distribution amount and taxability (e.g. Code 1/2/7)."],
        ["IRS Transcript", "Medicare Wages (Box 5) vs Box 1 Wages", "Pull 'Wage & Income Transcripts' from IRS.gov (10-yr history online)", "Medicare wages serve as a proxy for state-taxable wages if W-2 is lost."],
        ["Broker Ledger", "Participant Lifetime Contribution History", "Request from custodian customer service", "Identifies lifetime pre-tax employee deferrals (100% basis in NJ/PA)."]
    ]
    
    for row_offset, row_data in enumerate(checklist_data):
        row_idx = 14 + row_offset
        ws_inst.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_inst.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = data_border
            # Zebra striping
            if row_idx % 2 == 1:
                cell.fill = stripe_fill
                
    # ----------------------------------------------------
    # Sheet 2: NJ & MA Pro-Rata (Worksheet C)
    # ----------------------------------------------------
    ws_nj = wb.create_sheet(title="NJ & MA Pro-Rata")
    ws_nj.views.sheetView[0].showGridLines = True
    
    ws_nj["A1"] = "New Jersey & Massachusetts Pro-Rata Basis Roll-Forward (Worksheet C)"
    ws_nj["A1"].font = title_font
    ws_nj.row_dimensions[1].height = 30
    
    ws_nj["A3"] = "Formula: Excludable Portion = Distribution * (Basis / (Dec 31 Balance + Distribution))"
    ws_nj["A3"].font = italic_help_font
    
    nj_headers = [
        "Year", "Starting Basis", "Annual Contributions", 
        "Basis Before Withdrawal", "Total Distributions", "Dec 31 Account Value", 
        "Excludable Portion (Basis Return)", "Taxable Portion", "Ending Basis"
    ]
    
    for col_idx, h in enumerate(nj_headers, start=1):
        cell = ws_nj.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx == 1 else align_right
        cell.border = data_border
    ws_nj.row_dimensions[5].height = 28
    
    start_year = 2010
    total_years = 16 # 2010 to 2025
    
    for i in range(total_years):
        row_idx = 6 + i
        ws_nj.row_dimensions[row_idx].height = 22
        
        # Year
        cell_yr = ws_nj.cell(row=row_idx, column=1, value=start_year + i)
        cell_yr.font = bold_data_font
        cell_yr.alignment = align_center
        cell_yr.border = data_border
        
        # Starting Basis (Formula linked to previous Ending Basis, except first year)
        cell_start = ws_nj.cell(row=row_idx, column=2)
        if i == 0:
            cell_start.value = 0.00
        else:
            cell_start.value = f"=I{row_idx-1}"
        cell_start.font = data_font
        cell_start.number_format = "$#,##0.00"
        cell_start.border = data_border
        
        # Annual Contributions (Manual Entry)
        cell_contrib = ws_nj.cell(row=row_idx, column=3, value=0.00)
        cell_contrib.font = data_font
        cell_contrib.number_format = "$#,##0.00"
        cell_contrib.border = data_border
        
        # Basis Before Withdrawal (Formula: Starting Basis + Contributions)
        cell_before = ws_nj.cell(row=row_idx, column=4, value=f"=B{row_idx}+C{row_idx}")
        cell_before.font = data_font
        cell_before.number_format = "$#,##0.00"
        cell_before.border = data_border
        
        # Total Distributions (Manual Entry)
        cell_dist = ws_nj.cell(row=row_idx, column=5, value=0.00)
        cell_dist.font = data_font
        cell_dist.number_format = "$#,##0.00"
        cell_dist.border = data_border
        
        # Dec 31 Account Value (Manual Entry)
        cell_val = ws_nj.cell(row=row_idx, column=6, value=0.00)
        cell_val.font = data_font
        cell_val.number_format = "$#,##0.00"
        cell_val.border = data_border
        
        # Excludable Portion (Formula: Distribution * Min(BasisBefore, TotalVal) / TotalVal)
        # Formula uses Excel's MIN and IF to handle zero-division cleanly
        cell_excl = ws_nj.cell(row=row_idx, column=7, value=f"=IF((F{row_idx}+E{row_idx})>0, ROUND(E{row_idx}*MIN(D{row_idx}, F{row_idx}+E{row_idx})/(F{row_idx}+E{row_idx}), 2), 0)")
        cell_excl.font = bold_data_font
        cell_excl.fill = alert_fill
        cell_excl.number_format = "$#,##0.00"
        cell_excl.border = data_border
        
        # Taxable Portion (Formula: Distribution - Excludable)
        cell_tax = ws_nj.cell(row=row_idx, column=8, value=f"=MAX(0, E{row_idx}-G{row_idx})")
        cell_tax.font = data_font
        cell_tax.number_format = "$#,##0.00"
        cell_tax.border = data_border
        
        # Ending Basis (Formula: BasisBefore - Excludable)
        cell_end = ws_nj.cell(row=row_idx, column=9, value=f"=D{row_idx}-G{row_idx}")
        cell_end.font = bold_data_font
        cell_end.number_format = "$#,##0.00"
        cell_end.border = data_border
        
        # Zebra striping
        if row_idx % 2 == 1:
            for c in range(2, 7):
                ws_nj.cell(row=row_idx, column=c).fill = stripe_fill
            ws_nj.cell(row=row_idx, column=8).fill = stripe_fill
            
    # ----------------------------------------------------
    # Sheet 3: PA Cost Recovery
    # ----------------------------------------------------
    ws_pa = wb.create_sheet(title="PA Cost Recovery")
    ws_pa.views.sheetView[0].showGridLines = True
    
    ws_pa["A1"] = "Pennsylvania Cost Recovery Basis Roll-Forward (FIFO)"
    ws_pa["A1"].font = title_font
    ws_pa.row_dimensions[1].height = 30
    
    ws_pa["A3"] = "Formula: Excludable Portion = Min(Basis Before Withdrawal, Distribution) [Qualified Distributions at 59.5+ are 100% Tax-Free]"
    ws_pa["A3"].font = italic_help_font
    
    pa_headers = [
        "Year", "Starting Basis", "Annual Contributions", 
        "Basis Before Withdrawal", "Total Distributions", 
        "Excludable Portion (FIFO Basis Return)", "Taxable Portion", "Ending Basis"
    ]
    
    for col_idx, h in enumerate(pa_headers, start=1):
        cell = ws_pa.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx == 1 else align_right
        cell.border = data_border
    ws_pa.row_dimensions[5].height = 28
    
    for i in range(total_years):
        row_idx = 6 + i
        ws_pa.row_dimensions[row_idx].height = 22
        
        # Year
        cell_yr = ws_pa.cell(row=row_idx, column=1, value=start_year + i)
        cell_yr.font = bold_data_font
        cell_yr.alignment = align_center
        cell_yr.border = data_border
        
        # Starting Basis
        cell_start = ws_pa.cell(row=row_idx, column=2)
        if i == 0:
            cell_start.value = 0.00
        else:
            cell_start.value = f"=H{row_idx-1}"
        cell_start.font = data_font
        cell_start.number_format = "$#,##0.00"
        cell_start.border = data_border
        
        # Annual Contributions
        cell_contrib = ws_pa.cell(row=row_idx, column=3, value=0.00)
        cell_contrib.font = data_font
        cell_contrib.number_format = "$#,##0.00"
        cell_contrib.border = data_border
        
        # Basis Before Withdrawal
        cell_before = ws_pa.cell(row=row_idx, column=4, value=f"=B{row_idx}+C{row_idx}")
        cell_before.font = data_font
        cell_before.number_format = "$#,##0.00"
        cell_before.border = data_border
        
        # Total Distributions
        cell_dist = ws_pa.cell(row=row_idx, column=5, value=0.00)
        cell_dist.font = data_font
        cell_dist.number_format = "$#,##0.00"
        cell_dist.border = data_border
        
        # Excludable Portion (Formula: Min(BasisBefore, Distribution))
        cell_excl = ws_pa.cell(row=row_idx, column=6, value=f"=ROUND(MIN(D{row_idx}, E{row_idx}), 2)")
        cell_excl.font = bold_data_font
        cell_excl.fill = alert_fill
        cell_excl.number_format = "$#,##0.00"
        cell_excl.border = data_border
        
        # Taxable Portion (Formula: Distribution - Excludable)
        cell_tax = ws_pa.cell(row=row_idx, column=7, value=f"=MAX(0, E{row_idx}-F{row_idx})")
        cell_tax.font = data_font
        cell_tax.number_format = "$#,##0.00"
        cell_tax.border = data_border
        
        # Ending Basis (Formula: BasisBefore - Excludable)
        cell_end = ws_pa.cell(row=row_idx, column=8, value=f"=D{row_idx}-F{row_idx}")
        cell_end.font = bold_data_font
        cell_end.number_format = "$#,##0.00"
        cell_end.border = data_border
        
        # Zebra striping
        if row_idx % 2 == 1:
            for c in range(2, 5):
                ws_pa.cell(row=row_idx, column=c).fill = stripe_fill
            ws_pa.cell(row=row_idx, column=7).fill = stripe_fill
            
    # ----------------------------------------------------
    # Adjust Column Widths Dynamically
    # ----------------------------------------------------
    for ws in [ws_inst, ws_nj, ws_pa]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip title cells when evaluating widths to avoid massive columns
                if cell.row in [1, 3] and ws in [ws_nj, ws_pa]:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Hardcode a few specific wide columns in the instruction sheet
    ws_inst.column_dimensions["A"].width = 25
    ws_inst.column_dimensions["B"].width = 30
    ws_inst.column_dimensions["C"].width = 50
    ws_inst.column_dimensions["D"].width = 60
    
    # Save Workbook
    wb.save(filename)
    print(f"Excel basis recovery sheet successfully generated: {filename}")

if __name__ == "__main__":
    create_basis_sheet()
